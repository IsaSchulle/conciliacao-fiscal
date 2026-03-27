"""
Conciliação Fiscal - SEFAZ SC x Sistema
Versão Web (Streamlit)
"""

import streamlit as st
import pandas as pd
import re
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ─────────────────────────────────────────────
#  CONFIGURAÇÃO
# ─────────────────────────────────────────────
ESPECIES_FRETE          = {"CTE", "CTEOS"}
ESPECIES_FATURA         = {"NF", "NFCOM", "DANF3E", "NF3E"}
CFOPS_CONSUMO_IMOBILIZADO = {"1551", "2551", "1556", "2556", "1407", "2407"}
CFOPS_SEM_CREDITO       = {"1653", "2653", "1911", "2911", "1403", "2403", "1912", "2912"}
CFOPS_SEM_ICMS          = CFOPS_CONSUMO_IMOBILIZADO | CFOPS_SEM_CREDITO

COR_HEADER      = "1F4E79"
COR_HEADER_FONT = "FFFFFF"
COR_OK          = "C6EFCE"
COR_DIV         = "FFCCCC"
COR_SEM_SAT     = "FFF2CC"
COR_SEM_SIS     = "DDEBF7"
COR_ZEBRA       = "F2F2F2"

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def limpar_chave(valor):
    if pd.isna(valor):
        return None
    return re.sub(r"[^0-9]", "", str(valor))

def classificar_especie(especie):
    e = str(especie).strip().upper()
    if e in ESPECIES_FRETE:
        return "Frete"
    if e in ESPECIES_FATURA:
        return "Fatura"
    return "NFE"

def fmt_data(valor):
    if pd.isna(valor):
        return ""
    s = str(valor).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        try:
            return pd.to_datetime(s).strftime("%d/%m/%Y")
        except Exception:
            pass
    try:
        return pd.to_datetime(s, dayfirst=True).strftime("%d/%m/%Y")
    except Exception:
        return s

def normalizar_serie(valor):
    if pd.isna(valor):
        return ""
    return str(valor).strip().lstrip("0") or "0"

def fmt_valor(valor):
    try:
        return float(valor)
    except Exception:
        return 0.0

def safe_round(a, b):
    try:
        return round(float(a) - float(b), 2)
    except Exception:
        return ""

# ─────────────────────────────────────────────
#  LEITURA E LIMPEZA
# ─────────────────────────────────────────────
def carregar_sat(file):
    df = pd.read_excel(file, dtype=str)
    df.columns = df.columns.str.strip()
    df["_chave"] = df["ChaveAcesso"].apply(limpar_chave)
    df = df.dropna(subset=["_chave"])
    df = df[df["_chave"].str.len() == 44]
    return df

def carregar_sistema(file):
    df = pd.read_excel(file, sheet_name=0, dtype=str)
    df.columns = df.columns.str.strip()
    df["_chave"] = df["Chave de Acesso"].apply(limpar_chave)
    df["_tipo"] = df["Espécie"].apply(classificar_especie)
    df.loc[df["_chave"].notna() & (df["_chave"].str.len() != 44), "_chave"] = None

    for col_orig, col_soma in [
        ("Valor Contábil", "Soma_ValorContabil"),
        ("Base Cálculo",   "Soma_BC"),
        ("Valor Imposto",  "Soma_ICMS"),
    ]:
        df[col_orig] = pd.to_numeric(df[col_orig], errors="coerce").fillna(0)
        soma = df.groupby("_chave")[col_orig].transform("sum")
        df[col_soma] = soma

    return df

# ─────────────────────────────────────────────
#  CRUZAMENTO
# ─────────────────────────────────────────────
def cruzar(df_sys, df_sat):
    sat_cols = {
        "_chave": "_chave",
        "Situacao": "SAT_Situacao",
        "TipoDeOperacaoEntradaOuSaida": "SAT_TipoOperacao",
        "DataEmissao": "SAT_DataEmissao",
        "SerieDocumento": "SAT_Serie",
        "ValorTotalNota": "SAT_ValorTotal",
        "ValorBaseCalculoICMS": "SAT_BC_ICMS",
        "ValorTotalICMS": "SAT_ICMS",
        "ValorIPI": "SAT_IPI",
        "UltimoEventoDestinatario": "SAT_Manifesto",
    }
    df_sat_red = df_sat[list(sat_cols.keys())].rename(columns=sat_cols)
    merged = df_sys.merge(df_sat_red, on="_chave", how="outer", indicator=True)

    def status(row):
        if row["_merge"] == "both":        return "✔ OK"
        elif row["_merge"] == "left_only": return "⚠ Só no Sistema"
        else:                              return "⚠ Só no SEFAZ"

    merged["Status"] = merged.apply(status, axis=1)

    merged["DIF_Série"] = merged.apply(
        lambda r: "" if r["_merge"] != "both" else
        ("OK" if normalizar_serie(r.get("Série","")) == normalizar_serie(r.get("SAT_Serie","")) else "DIVERGE"),
        axis=1
    )
    merged["DIF_DataEmissão"] = merged.apply(
        lambda r: "" if r["_merge"] != "both" else
        ("OK" if fmt_data(r.get("Data Emissão","")) == fmt_data(r.get("SAT_DataEmissao","")) else "DIVERGE"),
        axis=1
    )

    def cfop_consumo(row):
        return str(row.get("Natureza", "")).strip() in CFOPS_SEM_ICMS

    def label_cfop(row):
        cfop = str(row.get("Natureza", "")).strip()
        if cfop in CFOPS_CONSUMO_IMOBILIZADO: return "Consumo/Imob."
        if cfop in CFOPS_SEM_CREDITO:         return "Sem Crédito ICMS"
        return ""

    def tem_valor_sistema(row):
        return (
            fmt_valor(row.get("Base Cálculo", 0)) != 0 or
            fmt_valor(row.get("Valor Imposto", 0)) != 0 or
            fmt_valor(row.get("Valor IPI", 0)) != 0
        )

    def calc_dif_bc(row):
        if row["_merge"] != "both": return ""
        if cfop_consumo(row):
            return "ERRO CFOP" if tem_valor_sistema(row) else label_cfop(row)
        return safe_round(row.get("Soma_BC", 0), row.get("SAT_BC_ICMS", 0))

    def calc_dif_icms(row):
        if row["_merge"] != "both": return ""
        if cfop_consumo(row):
            return "ERRO CFOP" if tem_valor_sistema(row) else label_cfop(row)
        return safe_round(row.get("Soma_ICMS", 0), row.get("SAT_ICMS", 0))

    def calc_dif_ipi(row):
        if row["_merge"] != "both": return ""
        if cfop_consumo(row):
            return "ERRO CFOP" if tem_valor_sistema(row) else label_cfop(row)
        return safe_round(row.get("Valor IPI", 0), row.get("SAT_IPI", 0))

    merged["DIF_BC_ICMS"] = merged.apply(calc_dif_bc, axis=1)
    merged["DIF_ICMS"]    = merged.apply(calc_dif_icms, axis=1)
    merged["DIF_IPI"]     = merged.apply(calc_dif_ipi, axis=1)

    merged["DIF_ValorTotal"] = merged.apply(
        lambda r: "" if r["_merge"] != "both" else
        safe_round(r.get("Soma_ValorContabil", 0), r.get("SAT_ValorTotal", 0)),
        axis=1
    )

    def gerar_alerta(row):
        alertas = []
        if str(row.get("SAT_Situacao", "")).strip().lower() == "cancelado":
            alertas.append("Nota Cancelada")
        if str(row.get("SAT_Manifesto", "")).strip().lower() == "operacao nao realizada":
            alertas.append("Operação não Realizada")
        if str(row.get("SAT_TipoOperacao", "")).strip().upper() == "E":
            alertas.append("Entrada no SEFAZ")
        if cfop_consumo(row) and tem_valor_sistema(row):
            alertas.append(f"Valores indevidos em CFOP {label_cfop(row)}")
        return " | ".join(alertas) if alertas else ""

    merged["ALERTA"] = merged.apply(gerar_alerta, axis=1)
    return merged

# ─────────────────────────────────────────────
#  COLUNAS DE SAÍDA
# ─────────────────────────────────────────────
COLS_SAIDA = [
    "ALERTA", "Status", "Chave de Acesso", "_tipo", "Espécie", "Empresa", "Filial",
    "Fornecedor", "CNPJ/CPF/CNO", "Número", "Natureza",
    "Série", "SAT_Serie", "DIF_Série",
    "Data Emissão", "SAT_DataEmissao", "DIF_DataEmissão",
    "Valor Contábil", "Soma_ValorContabil", "SAT_ValorTotal", "DIF_ValorTotal",
    "Base Cálculo", "Soma_BC", "SAT_BC_ICMS", "DIF_BC_ICMS",
    "Alíquota", "Valor Imposto", "Soma_ICMS", "SAT_ICMS", "DIF_ICMS",
    "Valor IPI", "SAT_IPI", "DIF_IPI",
    "SAT_Situacao", "SAT_TipoOperacao", "SAT_Manifesto",
]

HEADERS = {
    "ALERTA": "⚠ Revisar", "Status": "Status", "Chave de Acesso": "Chave de Acesso",
    "_tipo": "Tipo", "Espécie": "Espécie", "Empresa": "Empresa", "Filial": "Filial",
    "Fornecedor": "Fornecedor", "CNPJ/CPF/CNO": "CNPJ/CPF/CNO", "Número": "Número NF",
    "Natureza": "CFOP",
    "Série": "Série (Sis)", "SAT_Serie": "Série (SAT)", "DIF_Série": "Dif Série",
    "Data Emissão": "Emissão (Sis)", "SAT_DataEmissao": "Emissão (SAT)", "DIF_DataEmissão": "Dif Emissão",
    "Valor Contábil": "Valor Total (Sis)", "Soma_ValorContabil": "Soma Valor Total",
    "SAT_ValorTotal": "Valor Total (SAT)", "DIF_ValorTotal": "Dif Valor Total",
    "Base Cálculo": "BC ICMS (Sis)", "Soma_BC": "Soma BC ICMS",
    "SAT_BC_ICMS": "BC ICMS (SAT)", "DIF_BC_ICMS": "Dif BC ICMS",
    "Alíquota": "Alíq ICMS", "Valor Imposto": "ICMS (Sis)", "Soma_ICMS": "Soma ICMS",
    "SAT_ICMS": "ICMS (SAT)", "DIF_ICMS": "Dif ICMS",
    "Valor IPI": "IPI (Sis)", "SAT_IPI": "IPI (SAT)", "DIF_IPI": "Dif IPI",
    "SAT_Situacao": "Situação SEFAZ", "SAT_TipoOperacao": "Entrada/Saída", "SAT_Manifesto": "Manifesto",
}

LARGURAS = {
    "ALERTA": 30, "Status": 18, "Chave de Acesso": 50, "_tipo": 10,
    "Espécie": 10, "Empresa": 9, "Filial": 7, "Fornecedor": 12,
    "CNPJ/CPF/CNO": 22, "Número": 12, "Natureza": 10,
    "Série": 10, "SAT_Serie": 10, "DIF_Série": 12,
    "Data Emissão": 14, "SAT_DataEmissao": 14, "DIF_DataEmissão": 14,
    "Valor Contábil": 16, "Soma_ValorContabil": 16, "SAT_ValorTotal": 16, "DIF_ValorTotal": 14,
    "Base Cálculo": 16, "Soma_BC": 14, "SAT_BC_ICMS": 16, "DIF_BC_ICMS": 14,
    "Alíquota": 10, "Valor Imposto": 14, "Soma_ICMS": 14, "SAT_ICMS": 14, "DIF_ICMS": 12,
    "Valor IPI": 12, "SAT_IPI": 12, "DIF_IPI": 12,
    "SAT_Situacao": 18, "SAT_TipoOperacao": 15, "SAT_Manifesto": 28,
}

# ─────────────────────────────────────────────
#  ESCRITA DO EXCEL
# ─────────────────────────────────────────────
def escrever_aba(ws, df, titulo):
    thin = Side(style="thin", color="CCCCCC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS_SAIDA))
    cel_titulo = ws.cell(row=1, column=1, value=titulo)
    cel_titulo.font = Font(bold=True, size=13, color=COR_HEADER_FONT)
    cel_titulo.fill = PatternFill("solid", fgColor=COR_HEADER)
    cel_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for col_idx, col in enumerate(COLS_SAIDA, 1):
        cel = ws.cell(row=2, column=col_idx, value=HEADERS.get(col, col))
        cel.font = Font(bold=True, color=COR_HEADER_FONT)
        cel.fill = PatternFill("solid", fgColor=COR_HEADER)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = borda
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A3"

    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        status = str(row.get("Status", ""))
        zebra  = row_idx % 2 == 0

        if "Só no Sistema" in status:
            fill_row = PatternFill("solid", fgColor=COR_SEM_SAT)
        elif "Só no SEFAZ" in status:
            fill_row = PatternFill("solid", fgColor=COR_SEM_SIS)
        elif zebra:
            fill_row = PatternFill("solid", fgColor=COR_ZEBRA)
        else:
            fill_row = None

        for col_idx, col in enumerate(COLS_SAIDA, 1):
            valor = row.get(col, "")
            if pd.isna(valor): valor = ""
            cel = ws.cell(row=row_idx, column=col_idx, value=valor)
            cel.border = borda
            cel.alignment = Alignment(vertical="center")
            if fill_row:
                cel.fill = fill_row

            if col.startswith("DIF_") and valor not in ("", "OK"):
                if valor in ("Consumo/Imob.", "Sem Crédito ICMS"):
                    cel.fill = PatternFill("solid", fgColor="D9D9D9")
                    cel.font = Font(italic=True, color="595959")
                elif valor in ("DIVERGE", "ERRO CFOP"):
                    cel.fill = PatternFill("solid", fgColor=COR_DIV)
                    cel.font = Font(bold=True)
                else:
                    try:
                        num = float(valor)
                        if abs(num) > 0.01:
                            cel.fill = PatternFill("solid", fgColor=COR_DIV)
                            cel.font = Font(bold=True)
                    except Exception:
                        pass

            if col == "Status" and "✔" in status:
                cel.fill = PatternFill("solid", fgColor=COR_OK)
            if col == "ALERTA" and valor:
                cel.fill = PatternFill("solid", fgColor="FF6600")
                cel.font = Font(bold=True, color="FFFFFF")

    for col_idx, col in enumerate(COLS_SAIDA, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = LARGURAS.get(col, 14)


def escrever_resumo(ws, stats):
    thin = Side(style="thin", color="CCCCCC")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    for w, letra in zip([30,14,14,14,14], ["A","B","C","D","E"]):
        ws.column_dimensions[letra].width = w

    headers = ["Categoria", "Total", "✔ OK", "⚠ Só no Sistema", "⚠ Só no SEFAZ"]
    for col, h in enumerate(headers, 1):
        cel = ws.cell(row=1, column=col, value=h)
        cel.font = Font(bold=True, color=COR_HEADER_FONT)
        cel.fill = PatternFill("solid", fgColor=COR_HEADER)
        cel.border = borda
        cel.alignment = Alignment(horizontal="center")

    for row, (categoria, dados) in enumerate(stats.items(), 2):
        ws.cell(row=row, column=1, value=categoria).border = borda
        for col, key in enumerate(["total", "ok", "so_sis", "so_sat"], 2):
            cel = ws.cell(row=row, column=col, value=dados[key])
            cel.border = borda
            cel.alignment = Alignment(horizontal="center")


def gerar_excel(df_sat, df_sis):
    merged = cruzar(df_sis, df_sat)
    nfe     = merged[merged["_tipo"] == "NFE"]
    fretes  = merged[merged["_tipo"] == "Frete"]
    faturas = merged[merged["_tipo"] == "Fatura"]

    def stats(df):
        return {
            "total":  len(df),
            "ok":     (df["Status"] == "✔ OK").sum(),
            "so_sis": (df["Status"] == "⚠ Só no Sistema").sum(),
            "so_sat": (df["Status"] == "⚠ Só no SEFAZ").sum(),
        }

    resumo_stats = {
        "NFE":     stats(nfe),
        "Fretes":  stats(fretes),
        "Faturas": stats(faturas),
        "TOTAL":   stats(merged),
    }

    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Resumo"
    escrever_resumo(ws_res, resumo_stats)
    escrever_aba(wb.create_sheet("NFE"),     nfe,     "NFE — Conciliação Sistema x SEFAZ SC")
    escrever_aba(wb.create_sheet("Fretes"),  fretes,  "Fretes (CTE / CTEOS) — Conciliação Sistema x SEFAZ SC")
    escrever_aba(wb.create_sheet("Faturas"), faturas, "Faturas (NF / NFCOM etc.) — Conciliação Sistema x SEFAZ SC")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, resumo_stats

# ─────────────────────────────────────────────
#  INTERFACE STREAMLIT
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Conciliação Fiscal",
    page_icon="📊",
    layout="centered"
)

st.markdown("""
    <style>
        .block-container { max-width: 720px; }
        .stButton > button {
            background-color: #1F4E79;
            color: white;
            font-weight: bold;
            border-radius: 6px;
            padding: 0.5rem 2rem;
            width: 100%;
        }
        .stButton > button:hover { background-color: #163d61; }
        .resumo-card {
            background: #f0f4f8;
            border-radius: 8px;
            padding: 12px 20px;
            margin: 6px 0;
            font-size: 15px;
        }
    </style>
""", unsafe_allow_html=True)

st.markdown("## 📊 Conciliação Fiscal")
st.markdown("**SEFAZ SC × Sistema** — faça upload das duas planilhas e baixe o resultado pronto.")
st.divider()

col1, col2 = st.columns(2)
with col1:
    arquivo_sat = st.file_uploader("📄 Planilha SEFAZ SC (SAT)", type=["xlsx","xls"])
with col2:
    arquivo_sis = st.file_uploader("📄 Planilha do Sistema", type=["xlsx","xls"])

st.divider()

if st.button("▶  Gerar Conciliação"):
    if not arquivo_sat:
        st.error("Selecione a planilha do SEFAZ SC.")
    elif not arquivo_sis:
        st.error("Selecione a planilha do Sistema.")
    else:
        with st.spinner("Processando... aguarde"):
            try:
                df_sat = carregar_sat(arquivo_sat)
                df_sis = carregar_sistema(arquivo_sis)
                buf, resumo = gerar_excel(df_sat, df_sis)

                st.success("✅ Conciliação gerada com sucesso!")
                st.divider()
                st.markdown("#### 📋 Resumo")

                for cat, d in resumo.items():
                    emoji = "📦" if cat == "NFE" else "🚚" if cat == "Fretes" else "🧾" if cat == "Faturas" else "📊"
                    st.markdown(f"""
                        <div class='resumo-card'>
                            {emoji} <b>{cat}</b> &nbsp;|&nbsp;
                            Total: <b>{d['total']}</b> &nbsp;|&nbsp;
                            ✔ OK: <b>{d['ok']}</b> &nbsp;|&nbsp;
                            ⚠ Só Sistema: <b>{d['so_sis']}</b> &nbsp;|&nbsp;
                            ⚠ Só SEFAZ: <b>{d['so_sat']}</b>
                        </div>
                    """, unsafe_allow_html=True)

                st.divider()
                nome_arquivo = f"Conciliacao_{datetime.today().strftime('%Y%m%d')}.xlsx"
                st.download_button(
                    label="⬇️  Baixar Planilha de Conciliação",
                    data=buf,
                    file_name=nome_arquivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Erro ao processar: {e}")
